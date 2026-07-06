#!/usr/bin/env python3
"""Quick script to inspect the Excel file structure."""

import openpyxl

wb = openpyxl.load_workbook('data-eng/data/agricultural_indicator_schema_catalog.xlsx')
print('Sheet names:', wb.sheetnames)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f'\n=== {sheet_name} ===')
    print(f'Dimensions: {ws.max_row} rows x {ws.max_column} cols')
    print('First row (headers):', [cell.value for cell in ws[1]])
    if ws.max_row > 1:
        print('Second row (sample):', [cell.value for cell in ws[2]])
